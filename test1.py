import datetime
import time
import openpyxl
#import clipboard
import pyperclip
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys



html_file_path = 'UPdate.html'
with open(html_file_path, 'r', encoding='utf-8') as file:
    html_content = file.read()
pyperclip.copy(html_content)

path =f"MAAIL"

def login(Email, Password, emails):
    chrome_options = uc.ChromeOptions()
    #chrome_options.add_argument('--headless')  # Run Chrome in headless mode
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--window-position=0,400') 
    
    chrome_options.add_argument(f"--load-extension={path}")
    
    driver = uc.Chrome(options=chrome_options)
    driver.delete_all_cookies()
    


    # url1 = "https://chromewebstore.google.com/detail/html-inserter-for-gmail-o/omojcahabhafmagldeheegggbakefhlh?hl=en"
    # driver.get(url1)
    
    # WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Add to Chrome']"))).click()
    # time.sleep(2)
    # pyautogui.press('left')
    # time.sleep(1)
    # pyautogui.press('enter')

    #time.sleep(4)

   

   # time.sleep(4)

    url = "https://mail.google.com/mail/u/0/?tab=rm&ogbl#inbox"
    driver.get(url)
    driver.set_window_size(900,900)
    #driver.maximize_window()
    #driver.minimize_window()

    try:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[text()='Sign in']"))).click()
        time.sleep(2)
    except:
        pass

    try:
        email_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "identifierId")))
        email_input.send_keys(Email + "\n")
        time.sleep(5)
        password_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@type='password']")))
        password_input.send_keys(Password + "\n")

        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//span[normalize-space()='Not now']"))).click()
            time.sleep(3)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//span[normalize-space()='Not now']"))).click()
        except:
            pass
        
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@tabindex='0' and @role='button' and @class='bBe']"))).click()
        except:
            pass
        
        time.sleep(3)
        compose = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@role='button' and @gh='cm']")))
        compose.click()
        time.sleep(2)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "(//span[text()='Bcc'])[1]"))).click()
        time.sleep(2)
        to_input_mail = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@aria-label='BCC recipients']")))
        to_input_mail.send_keys(emails)
        time.sleep(1)
        
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//input[@name='subjectbox' and @class='aoT']"))).send_keys("You got a money request")
        time.sleep(1)
        
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Insert HTML']"))).click()
            time.sleep(2)
        except:
            print("can't find")     
        
        insrt= WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "(//div[@role='textbox'])[2]")))
        insrt.click()
        insrt.send_keys(Keys.CONTROL + 'a')  # Select all text
        time.sleep(1)
        insrt.send_keys(Keys.BACKSPACE)
        time.sleep(2)
       # insrt.send_keys(html_content)
        insrt.send_keys(Keys.CONTROL, 'v')
        
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[@id=':r1:']"))).click()
        time.sleep(3)
        send_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[text() = 'Send']")))
        send_button.click()
        time.sleep(3)
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        driver.quit()

wb = openpyxl.load_workbook('Forwading_Email.xlsx')
ws = wb.active

email_password_pairs = [(row[0], row[1]) for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)]

with open('Email.txt', 'r') as file:
    emails = [line.strip() for line in file.readlines()]

current_date = datetime.datetime.now()
if current_date < datetime.datetime(2024, 5, 22):
    batch_size = 5
    for i in range(0, len(emails), batch_size):
        batch_emails = emails[i:i + batch_size]
        for email_index, (email, password) in enumerate(email_password_pairs):
            if not batch_emails:
                break
            try:
                login(email, password, ', '.join(batch_emails))
                batch_emails = []
                email_password_pairs = email_password_pairs[email_index + 1:] + email_password_pairs[:email_index + 1]
            except Exception as e:
                print(f"Error forwarding emails with {email}: {e}")
                break
else:
    print("Please contact with Razib Khan")
